# -*- coding: utf-8 -*-

import os
import logging
import argparse

import pandas as pd
import numpy as np

import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill, numbers
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.alignment import Alignment
from openpyxl.formatting.rule import Rule
from datetime import date, timedelta
from datetime import datetime as dt
from dateutil import relativedelta
import copy

from dotenv import load_dotenv
load_dotenv()
print(os.environ)

FILEPATH = os.environ.get('dsr_outputpath')
IMS_DT_INPUT = os.environ.get('dsr_dt_input_filename')
IMS_DT_OUTPUT = os.environ.get('dsr_dt_output_filename')
IMS_DT_RAW_ACTUALS = os.environ.get('dsr_actuals_filename')

def get_current_load_dates():
    df = pd.read_csv(os.path.join(FILEPATH, IMS_DT_RAW_ACTUALS))
    df = df[['distributorcode', 'year', 'period', 'load_date']]
    df['load_date'] = pd.to_datetime(df['load_date'])
    df = df.sort_values(by='load_date')
    df = df.drop_duplicates(keep='last')
    df = df[(df['year'] == dt.now().year) & (df['period'] == dt.now().month)]
    cut_start = dt(dt.now().year, dt.now().month, dt.now().day, 9, 30, 0)
    prev_day = dt.today() - timedelta(days=1)
    cut_end = dt(prev_day.year, prev_day.month, prev_day.day, 9, 30, 0)
    df = df[df['load_date'].between(cut_end, cut_start)]
    df['load_date'] = df['load_date'] + timedelta(hours=4)
    df.rename(
            columns={
                'distributorcode': 'Distributor',
                'load_date': 'Load date (UAE)',
            },
            inplace=True
        )
    df.drop(
        ['year', 'period'],
        axis='columns',
        inplace=True
    )
    return df

def GetDRMIMSDT(dist, year, month):
    """
    Reads IMS DT csv and returns df
    """
    # load source
    df =  pd.read_csv(os.path.join(FILEPATH, IMS_DT_INPUT))
    #tidy up column names
    # mask negative values for btg_rr_required cols
    df['btg_rr_required'] = df['btg_rr_required'].mask(df['btg_rr_required'] < 0, 0)
    # filter selected distributor
    if dist is not None:
        df = df[df['distributorcode'].str.lower() == dist.lower()]
    # filter selected fyear and fmonth
    df = df[(df['year'] == year) & (df['period'] == f"P{month}")]
    # dropping unnecessary columns
    df.drop(
        ['year', 'quarter', 'period', 'y_remaining_days', 'q_remaining_days', 'month_days', 'days_gone', 'bu_region'],
        axis='columns',
        inplace=True
    )
    return df

def GenerateIMSDTExcel(fyear, fmonth, fdist):
    """
    Generates excel report for IMS DT
    """
    df_ims_dt = GetDRMIMSDT(fdist, fyear, fmonth)
    indexes = ['country_region', 'country', 'category', 'kelloggbrand']
    cols = [x for x in list(df_ims_dt.columns) if x not in indexes]
    cols.remove('distributorcode')
    agg_dct = {k:np.sum for k in cols}
    agg_dct['q_btg'] = 'first'
    agg_dct['y_btg'] = 'first'
    agg_dct['time_gone'] = 'first'
    
    ###innovation fltering
    df_innovation = df_ims_dt[df_ims_dt['kelloggbrand'].str.contains('INNOVATION')]
    df_kelloggbrand = df_ims_dt.groupby('kelloggbrand').agg(agg_dct).sort_index().reset_index()
    df_ims_dt = df_ims_dt[~(df_ims_dt['kelloggbrand'].str.contains('INNOVATION'))]
    df_innovation = pd.concat(
        [
            df_innovation.assign(
                **{x: '' for x in indexes[i:]}
            )
            .groupby(indexes)
            .agg(agg_dct) for i in range(1,6)
        ]
    ).sort_index()
    df_innovation.fillna(0, inplace=True)
    df_innovation = df_innovation.reset_index()
    custom_sort_cr = {
        "Middle East & Pakistan": 0,
        "SAUDI": 1,
        "GULF": 2,
        "LEVANT": 3,
        "PAKISTAN": 4,
    }
    custom_sort_c = {
        "": 0,
        "SAUDI": 1,
        "UAE": 2,
        "KUWAIT": 3,
        "OMAN": 4,
        "QATAR": 5,
        "BAHRAIN": 6,
        "IRAQ": 7,
        "JORDAN": 8,
        "LEBANON": 9,
        "PALESTINE": 10,
        "PAKISTAN": 11
    }

    df_innovation['country_region'] = pd.Categorical(df_innovation['country_region'], 
                         categories=sorted(custom_sort_cr, key=custom_sort_cr.get), 
                         ordered=True)
    df_innovation['country'] = pd.Categorical(df_innovation['country'], 
                         categories=sorted(custom_sort_c, key=custom_sort_c.get), 
                         ordered=True)
    df_innovation = df_innovation.sort_values(['country_region', 'country'])

    df_innovation.drop(
            ['country_region', 'category'],
            axis='columns',
            inplace=True
        )
    df_innovation = df_innovation[(df_innovation.country != '') & (df_innovation.kelloggbrand != '') ]
    ###
    #first grouping and aggregation
    df_output = pd.concat(
        [
            df_ims_dt.assign(
                **{x: '' for x in indexes[i:]}
            )
            .groupby(indexes)
            .agg(agg_dct) for i in range(1,6)
        ]
    ).sort_index()
    df_output.fillna(0, inplace=True)
    df_output = df_output.reset_index()

    df_output['country_region'] = pd.Categorical(df_output['country_region'], 
                         categories=sorted(custom_sort_cr, key=custom_sort_cr.get), 
                         ordered=True)
    df_output['country'] = pd.Categorical(df_output['country'], 
                         categories=sorted(custom_sort_c, key=custom_sort_c.get), 
                         ordered=True)
    df_output = df_output.sort_values(['country_region', 'country'])
    
    #grouping and agg for regions part
    df_regions = df_ims_dt.groupby('country_region').agg(agg_dct).sort_index().reset_index()
    
    df_mep = df_regions.sum(axis=0).to_frame()
    df_mep = df_mep.T
    df_regions = pd.concat([df_mep, df_regions[:]]).reset_index(drop = True)
    df_regions.loc[0, 'country_region'] = 'Middle East & Pakistan'
    df_regions.loc[0, 'time_gone'] = df_regions.loc[1, 'time_gone']

    df_regions = df_regions.iloc[df_regions['country_region'].map(custom_sort_cr).argsort()]
    
    df_category = df_ims_dt.groupby('category').agg(agg_dct).sort_index().reset_index()
    df_countries = df_ims_dt.groupby('country').agg(agg_dct).sort_index().reset_index()
    
    #flattening multiindex
    # df["Col3"], df["Col2"] = np.where(df['Col3'].isnull(), [df["Col2"], df["Col3"]], [df["Col3"], df["Col2"] ])
    for df_i in [df_output]:
        df_i['kelloggbrand'], df_i['category'] = np.where(df_i['kelloggbrand'] == '', [df_i['category'], df_i['kelloggbrand']], [df_i['kelloggbrand'], df_i['category'] ])
        df_i['kelloggbrand'], df_i['country'] = np.where(df_i['kelloggbrand'] == '', [df_i['country'], df_i['kelloggbrand']], [df_i['kelloggbrand'], df_i['country'] ])
        df_i['kelloggbrand'], df_i['country_region'] = np.where(df_i['kelloggbrand'] == '', [df_i['country_region'], df_i['kelloggbrand']], [df_i['kelloggbrand'], df_i['country_region'] ])
        df_i.drop(
            ['country_region', 'country', 'category'],
            axis='columns',
            inplace=True
        )
        df_i.rename(
            columns={
                'kelloggbrand': 'Country'
            },
            inplace=True
        )
    for df_i in [df_innovation]:
        df_i['kelloggbrand'], df_i['country'] = np.where(df_i['kelloggbrand'] == '', [df_i['country'], df_i['kelloggbrand']], [df_i['kelloggbrand'], df_i['country'] ])
        df_i.rename(
            columns={
                'kelloggbrand': 'Country'
            },
            inplace=True
        )
    # data sanitation and additional computing
    for df_i in [df_regions, df_category, df_output, df_innovation]:

        df_i['y_btg'] = (df_i['ytd_target_volume'] - df_i['ytd_achived_volume']) / df_i['ytd_target_volume'] * 100
        df_i['q_btg'] = (df_i['quarter_target_volume'] - df_i['qtd_achived_volume']) / df_i['quarter_target_volume'] * 100
        df_i['mtd_achived'] = np.true_divide(df_i['target_volume'], df_i['month_target']) * 100
        df_i['btg_volume'] = (df_i['month_target'] - df_i['target_volume'])
        df_i['btg_volume'] = df_i['btg_volume'].mask(df_i['btg_volume'] < 0, 0)
        df_i.replace([np.inf, -np.inf], 0, inplace=True)
        
        df_i['y_btg'] = df_i['y_btg'].map(lambda x: "{0:.2f}%".format(x) if not np.isnan(x) else "0.00%")
        df_i['q_btg'] = df_i['q_btg'].map(lambda x: "{0:.2f}%".format(x) if not np.isnan(x) else "0.00%")
        df_i['time_gone'] = df_i['time_gone'].map(lambda x: "{0:.2f}%".format(x*100) if not np.isnan(x) else "0.00%")
        df_i['mtd_achived'] = df_i['mtd_achived'].map(lambda x: "{0:.2f}%".format(x) if not np.isnan(x) else "0.00%")
    
        df_i.rename(
            columns={
                'month_target': 'Month Target',
                'time_gone': 'Time Gone',
                'target_volume': 'MTD Sale Volume',
                'mtd_achived': 'MTD % Achieved',
                'btg_volume': 'BTG Volume',
                'btg_rr_required': 'BTG RR Required to Hit the Month',
                'previous_month_ims': 'Previous Month IMS',
                'mtd_previous_month': 'MTD Previous Month',
                'quarter_target_volume': 'Quarter Target',
                'qtd_achived_volume': 'QTD Achieved',
                'q_btg': 'BTG Quarter %',
                'ytd_target_volume': 'Year Target',
                'ytd_achived_volume': 'YTD Achieved',
                'y_btg': 'BTG Year %'
            },
            inplace=True
        )
    df_regions.rename(
            columns={
                'country_region': 'Country Region',
            },
            inplace=True
        )
    
    # start of excel generation and formatting
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([''])
    ws.append([''])
    
    fill = PatternFill("solid", fgColor="F2F2F2")
    for r in dataframe_to_rows(df_regions, index=False, header=True):
        
        ws.append(r)
        regions_last_row = ws.max_row
        for i in range(0,len(df_output.columns)):
            cell = ws.cell(row=(regions_last_row),column=(i+1))
            cell.border = thin_border
            if i > 1:
                cell.alignment = Alignment(horizontal='right')
            if regions_last_row %2 == 0:
                cell.fill = fill
            if regions_last_row == 4:
                cell.font = Font(size="15", b=True)
    
    ws.append([''])

    for r in dataframe_to_rows(df_category, index=False, header=False):
        ws.append(r)
        category_last_row = ws.max_row
        for i in range(0,len(df_output.columns)):
            cell = ws.cell(row=(category_last_row),column=(i+1))
            cell.border = thin_border
            if i > 1:
                cell.alignment = Alignment(horizontal='right')
            if category_last_row %2 == 0:
                cell.fill = fill
    
    ws.append([''])

    output_last_row = 0
    inno_insert_positions = []
    for r in dataframe_to_rows(df_output, index=False, header=False):
        prev_row = [x.value for x in ws[ws.max_row]]
        if r != prev_row:
            if r[0] == "SAUDI" and output_last_row > 15:
                pass
            elif r[0] == "PAKISTAN" and output_last_row > 170:
                pass
            else:
                ws.append(r)
                output_last_row = ws.max_row
                if r[0] in list(df_countries['country']):
                    inno_insert_positions.append((r[0], output_last_row + 15))
                for i in range(0,len(df_output.columns)):
                    cell = ws.cell(row=(output_last_row),column=(i+1))
                    cell.border = thin_border
                    if i > 1:
                        cell.alignment = Alignment(horizontal='right')
                    if output_last_row %2 == 0:
                        cell.fill = fill
    inno_insert_positions.sort(key=lambda a: a[1], reverse=True)
    
    inno_rows = []
    df_innovation.drop_duplicates(inplace=True)
    for r in dataframe_to_rows(df_innovation, index=False, header=False):
        inno_rows.append(r)

    for pos in inno_insert_positions:
        for row in inno_rows:
            if pos[0] == row[0]:
                row.pop(0)
                ws.insert_rows(pos[1])
                for i in range(0,len(row)):
                    cell = ws.cell(row=(pos[1]),column=(i+1))
                    cell.border = thin_border
                    cell.value = row[i]
                    if pos[1] %2 == 0:
                        cell.fill = fill
        ws.insert_rows(pos[1])
        cell = ws.cell(row=(pos[1]),column=(1))
        cell.border = thin_border
        cell.value = "INNOVATION"
    ws.append([''])
    header = NamedStyle(name="header") 
    header.fill = PatternFill("solid", fgColor="c00000") # Dark-Red
    header.font = Font(bold=True, size=20, color="ffffff")
    header.alignment = Alignment(horizontal='center', vertical='center')

    # dim_holder = DimensionHolder(worksheet=ws)
    # for col in range(ws.min_column, ws.max_column + 1):
    #     dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col)
    # ws.column_dimensions = dim_holder

    hed_lst = [f'$A1="{x}"' for x in ['Country Region', 'category', 'Country']]
    
    dirt_yellow_fill = PatternFill(bgColor="FFF2CC")
    dxf = DifferentialStyle(fill=dirt_yellow_fill)
    for x in hed_lst:
        r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
        r.formula = [x]
        ws.conditional_formatting.add(f"A1:{get_column_letter(ws.max_column)}{ws.max_row}", r)
    
    cnt_lst = [f'$A1="{x}"' for x in list(df_regions['Country Region'])]
    
    blue_fill = PatternFill(bgColor="b9d4ed")
    dxf = DifferentialStyle(fill=blue_fill, font=Font(size="13"))
    for x in cnt_lst:
        r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
        r.formula = [x]
        ws.conditional_formatting.add(f"A1:{get_column_letter(ws.max_column)}{ws.max_row}", r)

    ws.merge_cells(f'A1:{get_column_letter(ws.max_column)}2')
    ws["A1"].style = header
    def last_day_of_month(any_day):
        return (any_day.replace(day=1) + timedelta(days=31)).replace(day=1) - timedelta(days=1)
    
    current_date = dt.now().date()
    if current_date.month != fmonth or current_date.day == 1:
        p = last_day_of_month(date(fyear, fmonth, 1))
    else:
        p = current_date
    ws["A1"] = f"DAILY SALES REPORT ({p.strftime('%d %B %Y')})"

    for row in ws.iter_rows():
        for cell in row:
            alignment = copy.copy(cell.alignment)
            alignment.wrapText=True
            cell.alignment = alignment

    cat_lst = list(df_category['category'])
    cat_lst.append('INNOVATION')
    cat_lst = [f'$A1="{x}"' for x in cat_lst]

    lightskyblue_fill = PatternFill(bgColor="D9E1F2")
    dxf = DifferentialStyle(fill=lightskyblue_fill, alignment=Alignment(horizontal='left'), font=Font(size=11))
    dxf_r = DifferentialStyle(fill=lightskyblue_fill, alignment=Alignment(horizontal='right'), font=Font(size=11))
    for x in cat_lst:
        r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
        r.formula = [x]
        r_r = Rule(type="expression", dxf=dxf_r, stopIfTrue=True)
        r_r.formula = [x]

        ws.conditional_formatting.add(f"A1:A{ws.max_row}", r)
        ws.conditional_formatting.add(f"B1:{get_column_letter(ws.max_column)}{ws.max_row}", r_r)
    
    prod_lst = [f'$A1="{x}"' for x in list(df_kelloggbrand['kelloggbrand'])]
    dxf = DifferentialStyle(alignment=Alignment(horizontal='right'), font=Font(size=11))
    for x in prod_lst:
        r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
        r.formula = [x]
        
        ws.conditional_formatting.add(f"A1:{get_column_letter(ws.max_column)}{ws.max_row}", r)
    
    cntry_lst = [f'$A1="{x}"' for x in list(df_countries['country'])]
    # #print(prod_lst)
    lightskyblue_fill = PatternFill(bgColor="f0adc5")
    dxf = DifferentialStyle(fill=lightskyblue_fill, alignment=Alignment(horizontal='left'), font=Font(size=12))
    dxf_r = DifferentialStyle(fill=lightskyblue_fill, alignment=Alignment(horizontal='right'), font=Font(size=12))
    for x in cntry_lst:
        r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
        r.formula = [x]
        r_r = Rule(type="expression", dxf=dxf_r, stopIfTrue=True)
        r_r.formula = [x]
        #print(category_last_row)
        ws.conditional_formatting.add(f"A1:A{ws.max_row}", r)
        ws.conditional_formatting.add(f"B1:{get_column_letter(ws.max_column)}{ws.max_row}", r_r)
    
    ws.freeze_panes = "A4"

    for row in ws.iter_rows():
        if row[0].value in list(df_kelloggbrand['kelloggbrand']) and row[0].value != "PRINGLES":
            row[0].value = ' '.join(elem.capitalize() for elem in row[0].value.split())
    
    ws.column_dimensions["A"].width = 30
    cell = ws.cell(row=3,column=2)
    cell.alignment = Alignment(horizontal='right')
    def char_range(c1, c2):
        """Generates the characters from `c1` to `c2`, inclusive."""
        for c in range(ord(c1), ord(c2)+1):
            yield chr(c)
    for x in char_range('B', get_column_letter(ws.max_column)):
        ws.column_dimensions[x].width = 13
    
    col_number_lst = [2, 4, 6, 7, 8, 9, 10, 11, 13, 14]
    for r in range(1,ws.max_row):
        for col in col_number_lst:
            cell = ws.cell(row=(r), column=(col))
            cell.number_format = '#,##0.00'

    ws.append([''])
    ws.append(["* Column H & I will show similar values. As for system we have consolidated data for August month sales. This will show properly from October month as we have daily files for September in the system."])
    ws.append(["* Some countries Target was not defined in KSOP file for some Brands hence will show Zero against it."])
    ws.append(["* All innovations will be captured in separate section to show the performance. These innovations as also part of their main Brand to capture the performance against whole target."])
    ws.append(["* Sales captures n-2 date. N is today's date. In the coming month, need to give deadline to distributor to share the files till 12noon Dubai time (for the previous day data) in order to generate the report by 3pm same day."])
    ws.append([''])
    ws.append([''])
    ws.append(["UPLOAD LOG"])
    df_upload_log = get_current_load_dates()
    if not df_upload_log.empty:
        for r in dataframe_to_rows(df_upload_log, index=False, header=True):
            ws.append(r)
            ws.merge_cells(f'B{ws.max_row}:C{ws.max_row}')
            for i in range(0,len(df_upload_log.columns)+1):
                cell = ws.cell(row=(ws.max_row),column=(i+1))
                cell.border = thin_border
                if ws.max_row %2 == 0:
                    cell.fill = fill
    else:
        ws.append(["No new data uploaded since last report generation."])
    
    wb.save(os.path.join(FILEPATH, IMS_DT_OUTPUT))

now = dt.now()
if now.day == 1:
    prev_month = now + relativedelta.relativedelta(months=-1)
    GenerateIMSDTExcel(prev_month.year, prev_month.month, None)
else:
    GenerateIMSDTExcel(now.year, now.month, None)
