# -*- coding: utf-8 -*-
"""
Date       : 22/02/2021
Author     :  Neeraj
Description: This script generates a Formatted excel output which Provides metrics  IMS ,OS, ADJ, WOC , Arrivals in Pivot up format
Project    : MENAT Pathways - Phase 1
Region     : AMEA
"""
import openpyxl
import os
import logging
import pandas as pd
import argparse
import sys
from datetime import datetime
import re
import numpy as np
import ast
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.colors import Color
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

load_dotenv()
print(os.environ)


## file paths
month_mapping=os.environ.get('month_mapping')
filepath = os.environ.get('filepath')

my_parser = argparse.ArgumentParser(prog="IPRXLTRANS", description="Prepare formatted Pipeline output.")


#Reading the argument
my_parser.add_argument(
    "--inputFilePath",
    metavar="Input File Path",
    type=str,
    help="Path to the input file"
)

my_parser.add_argument(
    "--outputFilePath",
    metavar="Output File Path",
    type=str,
    help="Path to the output file"
)

# my_parser.add_argument(
#     "--outputFileName",
#     metavar="Output File Name",
#     type=str,
#     help="Name of the output file"
# )

my_parser.add_argument(
    "--logFilePath",
    metavar="Log File Path",
    type=str,
    help="Path to the log file"
)

args = my_parser.parse_args()

if not os.path.exists(args.inputFilePath):
    print("Input path does not exists")
    sys.exit(1)

if not os.path.exists(args.outputFilePath):
    print("Output directory does not exist. Creating output directories")
    print("Creating output directory")
    os.makedirs(args.outputFilePath)

class DaysMissingException(Exception):
    def __init__(self, message="Some values are missing for Days column in the input file"):
        self.message = message
        super().__init__(self.message)


print(f"Input path is {args.inputFilePath}")
print(f"Output path is {args.outputFilePath}")

### Excel Input file path
LOC = args.inputFilePath

## Static variables for Headers in formatted excel

INPUT_HEADER_NAMES = [
    'Year',
    'Month',
    'Period',
    'YearMonth',
    'DistributorName',
    'DFU',
    'DFUDescription',
    'Brand',
    'Category',
    'conversion_factor',
    'OS',
    'ARR',
    'IMS',
    'WOC',
    'ADJ',
    'flag',
    'Days'
]

MONTHS = [
    'Jan',
    'Feb',
    'Mar',
    'Apr',
    'May',
    'Jun',
    'Jul',
    'Aug',
    'Sep',
    'Oct',
    'Nov',
    'Dec'
]

MONTH_NAMES_TO_MONTH_NUMBER_MAPPING = {
    '01': 'Jan',
    '02': 'Feb',
    '03': 'Mar',
    '04': 'Apr',
    '05': 'May',
    '06': 'Jun',
    '07': 'Jul',
    '08': 'Aug',
    '09': 'Sep',
    '10': 'Oct',
    '11': 'Nov',
    '12': 'Dec'
}

KPI_NAMES = [
    'WOC',
    'OS',
    'IMS',
    'ARR',
    'ADJ'
]

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# blue_colour = Color(rgb="#373A65")


def get_values(cell):
    """
    Helper function to return value of cell
    :param cell: cell of the sheet
    :return: string value
    """
    try:
        return cell.value
    except Exception as e:
        print("Returning after exception")
        return cell


def populate_month_columns(df_input):
    """

    :param df_input:
    :return:
    """
    #dfu_desc = df_input.iloc[0]['DFUDescription']
    #year_month = df_input[df_input['DFUDescription'] == dfu_desc]['YearMonth']
    year_month=df_input['YearMonth'].drop_duplicates()
    year_month_mapping = {}
    for ym in year_month:
        month = str(ym)[-2:]
        year = str(ym)[:4]
        if year in year_month_mapping.keys():
            year_month_mapping[year].append(month)
        else:
            year_month_mapping[year] = [month]
    for year, months in year_month_mapping.items():
        m = year_month_mapping[year]
        m = list(set(m))
        m.sort()
        year_month_mapping[year] = m
    MONTH_COLUMNS = []
    for year in sorted(year_month_mapping):
        months = year_month_mapping[year]
        for m in months:
            MONTH_COLUMNS.append(f'{MONTH_NAMES_TO_MONTH_NUMBER_MAPPING[m]}-{year[-2:]}')
    return MONTH_COLUMNS


# Processing input file
print("Reading the input file")
input_workbook = openpyxl.load_workbook(LOC, data_only=True)
input_sheet = input_workbook['Sheet1']

col_indices = [n for n, cell in enumerate(list(input_sheet.rows)[0]) if cell.value in INPUT_HEADER_NAMES]

df_input = pd.DataFrame(input_sheet)

df_input.drop(0, inplace=True) # Dropping the headers

df_input.drop(df_input.columns.difference(col_indices), axis=1, inplace=True)

col_map = {key: value for (key, value) in zip(col_indices, INPUT_HEADER_NAMES)}

df_input.rename(columns=col_map, inplace=True)

df_input = (df_input.applymap(get_values))
df_input_zero = df_input.copy()

## reading the months mapping file
df_ipr_months = pd.read_csv(os.path.join(filepath,month_mapping))
#print(df_ipr_months)
## harding the metrics value to zero for dummy month records
df_input_zero = df_input.drop(['YearMonth'],axis=1)
df_input_zero['conversion_factor'] = 0
df_input_zero['OS'] = 0
df_input_zero['ARR'] = 0
df_input_zero['IMS'] = 0
df_input_zero['WOC'] = 0
df_input_zero['ADJ'] = 0
df_input_zero['flag'] = 0
df_input_zero['Days'] = 0
#print(df_input_zero.count())
df_input_zero = df_input_zero.groupby(['Year','Month','Period','DistributorName','DFU',
                                      'DFUDescription','Brand','Category'])['conversion_factor','OS','ARR','IMS','WOC','ADJ','flag','Days'].max().reset_index()
   
#print(df_input_zero.count())
df_input_zero['key'] = 0
df_ipr_months['key'] = 0
df_input_cross_join = df_input_zero.merge(df_ipr_months,on='key')
#print(df_input_cross_join.count())
#df_input_cross_join.to_csv('test.csv',index=False,header=True)
df_input_cross_join.drop(['key'],axis=1,inplace=True)
df_input_cross_join = df_input_cross_join[['Year','Month','Period','YearMonth',
                                          'DistributorName','DFU','DFUDescription','Brand',
                                          'Category','conversion_factor','OS','ARR','IMS',
                                          'WOC','ADJ','flag','Days']]

#df_input_cross_join.to_csv('cross.csv')
print(df_input_cross_join.dtypes)
df_input_cross_join['YearMonth'] = df_input_cross_join['YearMonth'].astype(str)
#print(df_input_cross_join.count())
#df_input.drop(['conversion_factor'],axis=1,inplace=True)
#df_input_cross_join.drop(['conversion_factor'],axis=1,inplace=True)
df_input_final = df_input.append(df_input_cross_join,ignore_index=True)
#df_input_p = df_input_final.groupby(['Year','Month','Period','YearMonth','DistributorName','DFU',
#'DFUDescription','Brand','Category'],as_index=False)['OS', 'ARR', 'IMS', 'WOC', 'ADJ', 'flag', 'Days'].sum()
df_input_final['ARR'] = df_input_final.groupby(['Year','Month','Period','YearMonth','DistributorName','DFU',
'DFUDescription','Brand','Category'],as_index=False).ARR.transform(sum)
df_input_final[['IMS']] = df_input_final[['IMS']].replace(to_replace=[np.nan, 0], value=-999999999)  # Change 0's to respect negative values
df_input = df_input_final.groupby(['Year','Month','Period','YearMonth','DistributorName','DFU',
'DFUDescription','Brand','Category'],as_index=False)['conversion_factor', 'OS', 'ARR', 'IMS', 'WOC', 'ADJ', 'flag', 'Days'].max()
df_input.replace(to_replace=-999999999, value=0, inplace=True) #Revert 0's after max()
df_input['conversion_factor'] = df_input.groupby(['DistributorName','DFU'],as_index=False).conversion_factor.transform(max)
df_input.to_csv(os.path.join(filepath,'29_MPW_ExcelTransInputDebug.csv'))

null_number_of_days = df_input[df_input['Days'].isnull()]

if len(null_number_of_days):
    print("There are rows will missing Days values")
    logging.info("There are rows will missing Days values")
    raise DaysMissingException()


# year = df_input['Year'].iloc[0]
print('df_input now has aggregated YearMonth')
print(df_input.dtypes)
MONTH_COLUMNS = populate_month_columns(df_input)


# print(f'{df_input["Month"].iloc[0]}-{(df_input["Year"].iloc[0])[1:]}')
# print(df_input['Year'].iloc[0][-2:])
# print()
month = MONTH_COLUMNS.index(f'{df_input["Month"].iloc[0]}-{str(df_input["Year"].iloc[0])[-2:]}') + 1


# Processing output file


def create_headers(sheet):
    """
    Helper function to add a header to the sheet and return it
    :param sheet: The current active sheet.
    :return: sheet
    """
    print("Adding headers")
    #logging.info("Adding headers")
    print("Adding DFU Code")
    #logging.info("Adding DFU Code")
    # DFU Code
    cell = sheet.cell(row=2, column=1)
    cell.value = 'DFU CODE'
    cell.border = thin_border
    cell.font = Font(color="ffffff", bold=True)
    cell.fill = PatternFill("solid", fgColor="4D7CA2")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    cell = sheet.cell(row=3, column=1)
    cell.value = 'DFU CODE'
    cell.border = thin_border
    cell.font = Font(color="ffffff", bold=True)
    cell.fill = PatternFill("solid", fgColor="4D7CA2")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    print("Adding Description")
    #logging.info("Adding Description")
    # Description
    cell = sheet.cell(row=2, column=2)
    cell.value = 'DESCRIPTION'
    cell.border = thin_border
    cell.font = Font(color="ffffff", bold=True)
    cell.fill = PatternFill("solid", fgColor="4D7CA2")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    cell = sheet.cell(row=3, column=2)
    cell.value = 'DESCRIPTION'
    cell.border = thin_border
    cell.font = Font(color="ffffff", bold=True)
    cell.fill = PatternFill("solid", fgColor="4D7CA2")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    print("Adding Brand and Category")
    #logging.info("Adding Brand and Category")
    # Brand and Category
    cell = sheet.cell(row=2, column=3)
    cell.value = 'BRAND'
    cell.border = thin_border
    cell.font = Font(color="ffffff", bold=True)
    cell.fill = PatternFill("solid", fgColor="4D7CA2")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    cell = sheet.cell(row=3, column=3)
    cell.value = 'CATEGORY'
    cell.border = thin_border
    cell.font = Font(color="ffffff", bold=True)
    cell.fill = PatternFill("solid", fgColor="4D7CA2")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    print("Adding CS_WT")
    #logging.info("Adding CS_WT")
    # CS_WT
    cell = sheet.cell(row=2, column=4)
    cell.value = 'CS_WT'
    cell.border = thin_border
    cell.font = Font(color="ffffff", bold=True)
    cell.fill = PatternFill("solid", fgColor="4D7CA2")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    cell = sheet.cell(row=3, column=4)
    cell.value = 'CS_WT'
    cell.border = thin_border
    cell.font = Font(color="ffffff", bold=True)
    cell.fill = PatternFill("solid", fgColor="4D7CA2")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    print("Adding Months Columns")
    #logging.info("Adding Months Columns")
    # Merged year columns
    for (i, month) in zip(range(5, 5 * len(MONTH_COLUMNS) + 5, 5), MONTH_COLUMNS):
        cell = sheet.cell(row=2, column=i)
        cell.value = month
        cell.font = Font(color="ffffff", bold=True)
        cell.fill = PatternFill("solid", fgColor="696969")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell = sheet.cell(row=2, column=(i + 1))
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell = sheet.cell(row=2, column=(i + 2))
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell = sheet.cell(row=2, column=(i + 3))
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell = sheet.cell(row=2, column=(i + 4))
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells(start_row=2, end_row=2, start_column=i, end_column=(i + 4))
        print(f"{month} column added")
        #logging.info(f"{month} column added")

    print("Adding KPI columns")
    #logging.info("Adding KPI columns")
    for i in range(5, 5 * len(MONTH_COLUMNS) + 5, 5):
        cell = sheet.cell(row=3, column=i)
        cell.value = KPI_NAMES[0]
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(color="000000", bold=True)
        cell.fill = PatternFill("solid", fgColor="FFFF00")

        cell = sheet.cell(row=3, column=i + 1)
        cell.value = KPI_NAMES[1]
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(color="000000", bold=True)
        cell.fill = PatternFill("solid", fgColor="7CADD5")

        cell = sheet.cell(row=3, column=i + 2)
        cell.value = KPI_NAMES[2]
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(color="000000", bold=True)
        cell.fill = PatternFill("solid", fgColor="6BEA73")

        cell = sheet.cell(row=3, column=i + 3)
        cell.value = KPI_NAMES[3]
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(color="000000", bold=True)
        cell.fill = PatternFill("solid", fgColor="FAA641")

        cell = sheet.cell(row=3, column=i + 4)
        cell.value = KPI_NAMES[4]
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(color="000000", bold=True)
        cell.fill = PatternFill("solid", fgColor="F95831")

    print("KPIs added for all months")
    #logging.info("KPIs added for all months")
    return sheet


def get_dfu_code(dfu_description, df_input):
    """
    Helper function to get the DFU code
    :param dfu_description: The value of
     DFU Description.
    :param df_input: The dataframe of input
     file as created above.
    :return:
    """
    dfu_codes = df_input[df_input['DFUDescription'] == dfu_description]
    if len(dfu_codes) > 0:
        return dfu_codes['DFU'].iloc[0]
    else:
        return ''


def get_brand(dfu_description, df_input):
    """
    Helper function to get brand
    :param dfu_description: The value of
     DFU Description.
    :param df_input: The dataframe of input
     file as created above.
    :return:
    """
    brands = df_input[df_input['DFUDescription'] == dfu_description]
    if len(brands) > 0:
        return brands['Brand'].iloc[0]
    else:
        return ''


def fill_missing_values(value_list, length, character):
    """
    Helper function to fill missing values
    :param value_list: The list of KPI values
    :param length: The desired length of
     output KPI values
    :param character: The character to substitute
     in place of missing values
    :return:
    """
    offset = length - len(value_list)
    if offset:
        value_list.extend([character] * offset)
    return value_list


def get_cs_wt(dfu_description, df_input):
    """
    Helper function to get cs_wt value
    :param dfu_description: The value of
     DFU Description.
    :param df_input: The dataframe of input
     file as created above.
    :return:
    """
    cs_wt = df_input[df_input['DFUDescription'] == dfu_description]
    if len(cs_wt) > 0:
        return cs_wt['conversion_factor'].iloc[0]
    else:
        return None


def add_values(sheet, df_input):
    """
    Helper function to populate output sheet using input dataframe.
    :param sheet: Output sheet
    :param df_input: Input dataframe
    :return: sheet
    """
    df_input.drop_duplicates(subset=['DistributorName', 'DFU', 'DFUDescription', 'YearMonth', 'Brand'], keep="last", inplace=True)
    unique_dfu_desc = list(set(df_input['DFUDescription']))
    print(str(unique_dfu_desc))
    for dfu_desc in unique_dfu_desc:
        #dfu_desc=dfu_desc.encode('utf-8').decode('utf-8').strip()
        dfu_desc2=dfu_desc.encode('ascii', 'ignore').decode('ascii')
        last_row = sheet.max_row
        print(r"Processing {0} at row number {1}".format(dfu_desc2,last_row + 1))
        #logging.info(print(f"Processing {dfu_desc2} at row number {last_row + 1}"))
        if get_dfu_code(dfu_desc, df_input) is None or str(get_dfu_code(dfu_desc, df_input)) == 'nan':
            dfu_code = 'NA'
        else:
            dfu_code = get_dfu_code(dfu_desc, df_input)
        #logging.info(f"The dfu_code for {dfu_desc2} is {dfu_code}")
        print(f"The dfu_code for {dfu_desc2} is {dfu_code}")

        if get_brand(dfu_desc, df_input) is None or str(get_brand(dfu_desc, df_input)) == 'nan':
            brand = 'NA'
        else:
            brand = get_brand(dfu_desc, df_input)
        #logging.info(f"The brand for {dfu_desc2} is {brand}")
        print(f"The brand for {dfu_desc2} is {brand}")

        if get_cs_wt(dfu_desc, df_input):
            conversion_factor = get_cs_wt(dfu_desc, df_input)
        else:
            conversion_factor = None
        print(f"The conversion factor for {dfu_desc2} is {conversion_factor}")
        #logging.info(f"The conversion factor for {dfu_desc2} is {conversion_factor}")

        df_temp = df_input[(df_input['Brand'] == brand) & (df_input['DFU'] == dfu_code)]
        category = df_input[df_input['DFUDescription'] == dfu_desc].iloc[0]['Category']
        #days = df_input[df_input['DFUDescription'] == dfu_desc].iloc[0]['Days']
        days = df_input[(df_input['Category'] == category)]['Days'].max()
        print(f"The stock cover days for {dfu_desc2} is {days}")
        #logging.info(f"The days for {dfu_desc2} is {days}")

        # Add DFU Code
        cell = sheet.cell(row=(last_row + 1), column=1)
        cell.value = dfu_code
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        print(f"Added DFU code for {dfu_desc2} at location {cell.coordinate}")
        #logging.info(f"Added DFU code for {dfu_desc2} at location {cell.coordinate}")

        # Add description
        cell = sheet.cell(row=(last_row + 1), column=2)
        cell.value = dfu_desc
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        print(f"Added DFUDescription for {dfu_desc2} at location {cell.coordinate}")
        #logging.info(f"Added DFUDescription for {dfu_desc2} at location {cell.coordinate}")

        # Add Brand
        cell = sheet.cell(row=(last_row + 1), column=3)
        cell.value = brand
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        print(f"Added Brand for {dfu_desc2} at location {cell.coordinate}")
        #logging.info(f"Added Brand for {dfu_desc2} at location {cell.coordinate}")

        # Adding CS_WT
        cell = sheet.cell(row=(last_row + 1), column=4)
        cell.value = conversion_factor
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        print(f"Added CS_WT for {dfu_desc2} at location {cell.coordinate}")
        #logging.info(f"Added CS_WT for {dfu_desc2} at location {cell.coordinate}")

        # Adding KPI values
        value_list = []
        for i, row in df_temp.iterrows():
            if conversion_factor:
                value_list.extend([
                    row['WOC'],
                    row['OS'] * conversion_factor,
                    row['IMS'] * conversion_factor,
                    row['ARR'] * conversion_factor,
                    row['ADJ']
                ])
            else:
                value_list.extend([
                    row['WOC'],
                    row['OS'],
                    row['IMS'],
                    row['ARR'],
                    row['ADJ']
                ])
        value_list = fill_missing_values(value_list, 5 * len(MONTH_COLUMNS), '-')
        flag_temp = df_input[df_input['DFUDescription'] == dfu_desc]['flag']
        flag_list = []
        for f in flag_temp:
            flag_list.extend([f] * 5)
        print("Adding KPI values")
        for i, flag, value in zip(range(5, 5 * len(MONTH_COLUMNS) + 5), flag_list, value_list):
            print(i)
            # flag = df_input[df_input['DFUDescription'] == dfu_desc].iloc[0]['flag']
            print(f"The flag for {dfu_desc2} month {value} is {flag}")
            #logging.info(f"The flag for {dfu_desc2} month {value} is {flag}")
            if (i % 5) == 0:  # Formula for WOC # CELL 5
                # row['WOC'],
                # row['OS'],
                # row['IMS'],
                # row['ARR'],
                # row['ADJ']
                # cell = sheet.cell(row=(last_row + 1), column=i)
                # # cell.value = f'=({sheet.cell(row=(last_row + 1), column=i - 4).coordinate} + {sheet.cell(row=(last_row + 1), column=i - 2).coordinate} - {sheet.cell(row=(last_row + 1), column=i - 3).coordinate} - {sheet.cell(row=(last_row + 1), column=i + 1).coordinate})'
                # cell.value = f'=((({sheet.cell(row=(last_row + 1), column=i - 4).coordinate} + {sheet.cell(row=(last_row + 1), column=i - 1).coordinate}) - {sheet.cell(row=(last_row + 1), column=i - 2).coordinate}) - {sheet.cell(row=(last_row + 1), column=i + 2).coordinate})'
                # # print(f'=MAX((({sheet.cell(row=(last_row + 1), column=i - 4).coordinate} + {sheet.cell(row=(last_row + 1), column=i - 2).coordinate}) - {sheet.cell(row=(last_row + 1), column=i - 3).coordinate}) - {sheet.cell(row=(last_row + 1), column=i + 1).coordinate}, 0)')
                # cell.border = thin_border
                # cell.alignment = Alignment(horizontal="center", vertical="center")

                print(f"Printing for {month}, {flag}")
                cell = sheet.cell(row=(last_row + 1), column=i)
                # =IFERROR(((((AE6 + AJ6) / 61) * 60) - (Y6 - Z6)), "-")
                # Positive values
                # cell.value = f'=MAX(IFERROR((((({sheet.cell(row=(last_row + 1), column=i + 4).coordinate}+{sheet.cell(row=(last_row + 1), column=i + 9).coordinate})/61)*{days}) - ({sheet.cell(row=(last_row + 1), column=i - 2).coordinate} - {sheet.cell(row=(last_row + 1), column=i - 1).coordinate})),0), 0)'
                # All values
                cell.value = f'=IFERROR({sheet.cell(row=(last_row + 1), column=(i + 1)).coordinate}/({sheet.cell(row=(last_row + 1), column=(i + 2)).coordinate}+{sheet.cell(row=(last_row + 1), column=(i + 7)).coordinate})*(61),0)'  # noqa
                # cell.value = f'=MAX(IFERROR((((({sheet.cell(row=(last_row + 1), column=i + 4).coordinate}+{sheet.cell(row=(last_row + 1), column=i + 9).coordinate})/61)*{days}) - ({sheet.cell(row=(last_row + 1), column=i - 2).coordinate} - {sheet.cell(row=(last_row + 1), column=i - 1).coordinate})),0), 0)'
                # cell.value = f'=IFERROR({sheet.cell(row=(last_row + 1), column=i - 2).coordinate} - ({sheet.cell(row=(last_row + 1), column=(i + 4)).coordinate} + {sheet.cell(row=(last_row + 1), column=(i + 8)).coordinate}) / 61 * {days} - {sheet.cell(row=(last_row + 1), column=(i - 1)).coordinate}, "-")'  #noqa
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue
            if ((i - 1) % 5 == 0) and (i >= ((month + 1) * 5)):  # Formula for OS for future months # CELL 1
                cell = sheet.cell(row=(last_row + 1), column=i)
                # Positive Values
                # cell.value = f'=MAX(IFERROR({sheet.cell(row=(last_row + 1), column=(i - 3)).coordinate}/({sheet.cell(row=(last_row + 1), column=(i - 2)).coordinate}+{sheet.cell(row=(last_row + 1), column=(i + 3)).coordinate})*(61),0), 0)'  # noqa
                # All values
                cell = sheet.cell(row=(last_row + 1), column=i)
                cell.value = f'=MAX(0, {sheet.cell(row=(last_row + 1), column=i - 5).coordinate} + {sheet.cell(row=(last_row + 1), column=i - 3).coordinate} - {sheet.cell(row=(last_row + 1), column=i - 4).coordinate})'  # noqa
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue
                # cell.value = f'=IFERROR({sheet.cell(row=(last_row + 1), column=(i + 1)).coordinate}/({sheet.cell(row=(last_row + 1), column=(i + 2)).coordinate}+{sheet.cell(row=(last_row + 1), column=(i + 6)).coordinate})*(61),0)'  # noqa
                # cell.border = thin_border
                # cell.alignment = Alignment(horizontal="center", vertical="center")
                # continue
            if ((i - 3) % 5) == 0 and int(flag) == 1:  # Formula for ARR # CELL 2
                cell = sheet.cell(row=(last_row + 1), column=i)
                # cell.value = f'=MAX(0, {sheet.cell(row=(last_row + 1), column=i - 5).coordinate} + {sheet.cell(row=(last_row + 1), column=i - 4).coordinate} - {sheet.cell(row=(last_row + 1), column=i - 3).coordinate})'  # noqa
                cell.value = f'=MAX(IFERROR((((({sheet.cell(row=(last_row + 1), column=i + 4).coordinate}+{sheet.cell(row=(last_row + 1), column=i - 1).coordinate})/61)*{days}) - ({sheet.cell(row=(last_row + 1), column=i - 2).coordinate} - {sheet.cell(row=(last_row + 1), column=i + 9).coordinate})),0), 0)'
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue
            if (i -4) % 5 == 0:  # Formula for ADJ # CELL 4
                cell = sheet.cell(row=(last_row + 1), column=i)
                # row['WOC'],
                # row['OS'],
                # row['IMS'],
                # row['ARR'],
                # row['ADJ']
                # cell.value = f'=({sheet.cell(row=(last_row + 1), column=i - 4).coordinate} + {sheet.cell(row=(last_row + 1), column=i - 2).coordinate} - {sheet.cell(row=(last_row + 1), column=i - 3).coordinate} - {sheet.cell(row=(last_row + 1), column=i + 1).coordinate})'
                cell.value = f'=((({sheet.cell(row=(last_row + 1), column=i - 3).coordinate} + {sheet.cell(row=(last_row + 1), column=i - 1).coordinate}) - {sheet.cell(row=(last_row + 1), column=i - 2).coordinate}) - {sheet.cell(row=(last_row + 1), column=i + 2).coordinate})'
                # print(f'=MAX((({sheet.cell(row=(last_row + 1), column=i - 4).coordinate} + {sheet.cell(row=(last_row + 1), column=i - 2).coordinate}) - {sheet.cell(row=(last_row + 1), column=i - 3).coordinate}) - {sheet.cell(row=(last_row + 1), column=i + 1).coordinate}, 0)')
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue
            cell = sheet.cell(row=(last_row + 1), column=i)
            # Positive values
            # if value > 0:
            #     cell.value = value
            # else:
            #     cell.value = 0
            # All values
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    return sheet


distributor_names = df_input['DistributorName'].unique()
print(distributor_names)
## running for 1 distributor and creating formatted file
for distributor_name in distributor_names:
    df_temp = df_input[df_input['DistributorName'] == distributor_name]
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'IntegratedPipeline'
    sheet = create_headers(sheet)
    sheet = add_values(sheet, df_temp)

    print("Saving the output...")
    #logging.info("Saving the output...")
    os.chdir(os.path.dirname(args.outputFilePath))
    print(os.path.dirname(args.outputFilePath))
    file_name = re.sub(r"[^a-zA-Z0-9]+", ' ', distributor_name)
    wb.save(f'{file_name.replace(" ", "_")}.xlsx')
    print(f"Saved {distributor_name} Successfully")
    #logging.info(f"Saved {distributor_name} Successfully")
