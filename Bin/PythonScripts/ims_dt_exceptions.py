# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
import glob
import os

from datetime import date, timedelta

import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side

from dotenv import load_dotenv
load_dotenv()
print(os.environ)

FILEPATH = os.environ.get('dsr_outputpath')
IMS_DT_EXCEPTION_OUTPUT = os.environ.get('dsr_exceptions_filename')
IMS_DT_ACTUALS = os.environ.get('dsr_actuals_filename')
IMS_DT_MATERIAL = os.environ.get('dsr_material_filename')

df_actuals = pd.read_csv(os.path.join(FILEPATH, IMS_DT_ACTUALS))
df_material = pd.read_csv(os.path.join(FILEPATH, IMS_DT_MATERIAL))

df_difference = df_actuals[~df_actuals['kelloggskucode'].isin(df_material['sku'])]
df_difference.fillna("", inplace=True)
# indexes = df_difference[df_difference['kelloggskucode'].str.contains(pat='.*[.]0', regex=True)].index
# df_difference.drop(indexes, inplace=True)
df_difference = df_difference.drop_duplicates(subset=['kelloggskucode'])

df_difference.drop(
    ['kelloggskudescription', 'kelloggbrand',
       'packsizegrams', 'caseweightgrams', 'promo_non_promo', 'channel',
       'subchannel', 'salesincases', 'distributorgsv', 'year', 'period'],
    axis='columns',
    inplace=True
)

df_difference['load_date'] = pd.to_datetime(df_difference['load_date'])
df_difference['load_date'] = df_difference['load_date'] + timedelta(hours=4)
df_difference = df_difference[['kelloggskucode', 'distributorcode', 'load_date']]
df_difference.load_date = df_difference.load_date.astype(str)
df_difference.rename(
            columns={
                'kelloggskucode': 'SKU',
                'distributorcode': 'Distributor',
                'load_date': 'Load date (UAE)',
            },
            inplace=True
        )

wb = openpyxl.Workbook()
ws = wb.active

thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

fill = PatternFill("solid", fgColor="F2F2F2")
for r in dataframe_to_rows(df_difference, index=False, header=True):
    ws.append(r)
    for i in range(0,len(df_difference.columns)):
        cell = ws.cell(row=(ws.max_row),column=(i+1))
        cell.border = thin_border
        if ws.max_row %2 == 0:
            cell.fill = fill

ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 30
#IMS_DT_EXCEPTION_OUTPUT=IMS_DT_EXCEPTIONS.xlsx
wb.save(os.path.join(FILEPATH, IMS_DT_EXCEPTION_OUTPUT))
